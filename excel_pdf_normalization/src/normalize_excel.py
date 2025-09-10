#!/usr/bin/env python3
import argparse
import pandas as pd
from openpyxl import load_workbook

def read_and_normalize_xlsx(path: str, sheet: str) -> pd.DataFrame:
    wb = load_workbook(path)
    ws = wb[sheet]
    # Unhide rows and unmerge cells
    for rng in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(rng))
    for row_dim in ws.row_dimensions.values():
        row_dim.hidden = False
    from io import BytesIO
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    df = pd.read_excel(bio, sheet_name=sheet, header=0)
    df.columns = [str(c).strip().lower().replace(' ', '_') for c in df.columns]
    return df

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--input', required=True)
    ap.add_argument('--sheet', required=True)
    ap.add_argument('--output', required=True)
    args = ap.parse_args()
    df = read_and_normalize_xlsx(args.input, args.sheet)
    df.to_csv(args.output, index=False)
    print("Excel normalized")

if __name__ == '__main__':
    main()
